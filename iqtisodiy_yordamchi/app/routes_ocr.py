#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OCR Routes - Hujjatlarni skanerlash va matn ajratish API
"""

from flask import Blueprint, request, jsonify, session, send_file
from werkzeug.utils import secure_filename
from datetime import datetime
from functools import wraps
import os
from dotenv import load_dotenv
import logging
import json
from pathlib import Path

load_dotenv()
logger = logging.getLogger(__name__)

# OCR Processor import
try:
    USE_DEV_OCR = os.getenv('USE_DEV_OCR', 'True').lower() == 'true'
    if USE_DEV_OCR:
        from app.ocr_processor import OCRProcessorDev as OCRProcessor, DocumentScanner
    else:
        from app.ocr_processor import OCRProcessor, DocumentScanner
except Exception as e:
    logger.warning(f"OCR processor import error: {e}")
    OCRProcessor = None
    DocumentScanner = None

def init_ocr_routes(app, db, User, login_required):
    """OCR routes-ni qayd qilish"""
    
    ocr_bp = Blueprint('ocr', __name__, url_prefix='/api/ocr')
    
    # Allowed file extensions
    ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'pdf', 'bmp', 'tiff'}
    UPLOAD_FOLDER = 'uploads/documents'
    
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    
    def allowed_file(filename):
        """Fayl formatni tekshirish"""
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    def handle_file_upload(prefix=""):
        """Fayl yuklashni boshqaradigan dekorator"""
        def decorator(f):
            @wraps(f)
            def decorated_function(*args, **kwargs):
                if OCRProcessor is None or DocumentScanner is None:
                    return jsonify({'error': 'OCR processor available emas'}), 500
                
                if 'file' not in request.files:
                    return jsonify({'error': 'Fayl yo\'q'}), 400
                
                file = request.files['file']
                
                if file.filename == '':
                    return jsonify({'error': 'Fayl tanlanmagan'}), 400
                
                if not allowed_file(file.filename):
                    return jsonify({'error': 'Noto\'g\'ri fayl formati. Allowed: jpg, png, pdf, bmp'}), 400
                
                # Faylni saqlash
                filename = secure_filename(f"{prefix}{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                file_path = os.path.join(UPLOAD_FOLDER, filename)
                file.save(file_path)
                
                return f(file_path=file_path, original_filename=file.filename, *args, **kwargs)
            return decorated_function
        return decorator
    
    @ocr_bp.route('/extract-text', methods=['POST'])
    @login_required
    @handle_file_upload()
    def extract_text(file_path, original_filename):
        """
        Rasmdan matn ajratib olish
        
        Request:
        - Multipart form data: file (rasm fayli)
        - Optional: lang (til kodi, default: uzb+eng)
        
        Response:
        {
            "success": true,
            "text": "Ajratilgan matn...",
            "confidence": 0.95,
            "file": "upload_file_path"
        }
        """
        try:
            # OCR ni ishga tushirish
            ocr = OCRProcessor()
            lang = request.form.get('lang', 'uzb+eng')
            result = ocr.extract_text_from_image(file_path, lang=lang)
            
            logger.info(f"Matn ajratildi: {original_filename}")
            
            return jsonify(result), 200
        
        except Exception as e:
            logger.error(f"Extract text error: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @ocr_bp.route('/extract-invoice', methods=['POST'])
    @login_required
    @handle_file_upload(prefix="invoice_")
    def extract_invoice(file_path, original_filename):
        """
        Hisob-fakturadan ma'lumotlarni ajratib olish
        
        Response:
        {
            "success": true,
            "data": {
                "invoice_number": "HF-2026-001",
                "invoice_date": "2026-02-04",
                "customer_name": "ABC Corp",
                "total_amount": 5000000,
                "items": [...]
            }
        }
        """
        try:
            # OCR ni ishga tushirish
            ocr = OCRProcessor()
            result = ocr.extract_invoice_data(file_path)
            
            logger.info(f"Hisob-faktura ajratildi: {original_filename}")
            
            return jsonify(result), 200
        
        except Exception as e:
            logger.error(f"Extract invoice error: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @ocr_bp.route('/scan-and-save', methods=['POST'])
    @login_required
    @handle_file_upload(prefix="scanned_")
    def scan_and_save(file_path, original_filename):
        """
        Hujjatni skanerlash va saqqlash
        
        Response:
        {
            "success": true,
            "file": "saved_file_path",
            "data": {...}
        }
        """
        try:
            # Skanerlash va saqqlash
            scanner = DocumentScanner()
            result = scanner.scan_and_save(file_path)
            
            logger.info(f"Hujjat saqlandi: {original_filename}")
            
            return jsonify(result), 200
        
        except Exception as e:
            logger.error(f"Scan and save error: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @ocr_bp.route('/batch-process', methods=['POST'])
    @login_required
    def batch_process():
        """
        Bir nechta hujjatlarni bir paytda qayta ishlash
        
        Request:
        {
            "files": [file1, file2, ...]
        }
        
        Response:
        {
            "success": true,
            "total_processed": 5,
            "successful": 5,
            "failed": 0,
            "results": [...]
        }
        """
        try:
            if DocumentScanner is None:
                return jsonify({'error': 'Document scanner available emas'}), 500
            
            if 'files' not in request.files:
                return jsonify({'error': 'Fayllar yo\'q'}), 400
            
            files = request.files.getlist('files')
            
            if not files:
                return jsonify({'error': 'Hech qanday fayl tanlanmagan'}), 400
            
            # Temp papka yaratish
            temp_folder = os.path.join(UPLOAD_FOLDER, f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
            os.makedirs(temp_folder, exist_ok=True)
            
            # Fayllarni saqlash
            saved_files = []
            for file in files:
                if allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    file_path = os.path.join(temp_folder, filename)
                    file.save(file_path)
                    saved_files.append(file_path)
            
            if not saved_files:
                return jsonify({'error': 'Hech qanday valyut fayl o\'rnatilmadi'}), 400
            
            # Batch processing
            scanner = DocumentScanner()
            results = []
            for file_path in saved_files:
                result = scanner.ocr.extract_invoice_data(file_path)
                results.append(result)
            
            logger.info(f"Batch processing tugallandi: {len(results)} ta fayl")
            
            return jsonify({
                'success': True,
                'total_processed': len(results),
                'successful': sum(1 for r in results if r.get('success', False)),
                'failed': sum(1 for r in results if not r.get('success', False)),
                'results': results
            }), 200
        
        except Exception as e:
            logger.error(f"Batch process error: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @ocr_bp.route('/get-uploaded-files', methods=['GET'])
    @login_required
    def get_uploaded_files():
        """
        Yuklangan hujjatlarni ro'yxatini olish
        """
        try:
            files = []
            
            if os.path.exists(UPLOAD_FOLDER):
                for file in os.listdir(UPLOAD_FOLDER):
                    file_path = os.path.join(UPLOAD_FOLDER, file)
                    if os.path.isfile(file_path):
                        files.append({
                            'name': file,
                            'path': file_path,
                            'size': os.path.getsize(file_path),
                            'created': datetime.fromtimestamp(os.path.getctime(file_path)).isoformat()
                        })
            
            return jsonify({
                'success': True,
                'total': len(files),
                'files': files
            }), 200
        
        except Exception as e:
            logger.error(f"Get files error: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @ocr_bp.route('/delete-file/<filename>', methods=['DELETE'])
    @login_required
    def delete_file(filename):
        """
        Yuklangan hujjatni o'chirish
        """
        try:
            user = db.session.query(User).filter_by(id=session['user_id']).first()
            if not user or user.role.name not in ['admin', 'manager']:
                return jsonify({'error': 'Huquq yo\'q'}), 403
            
            file_path = os.path.join(UPLOAD_FOLDER, secure_filename(filename))
            
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Fayl o'chirildi: {filename}")
                return jsonify({'success': True, 'message': 'Fayl o\'chirildi'}), 200
            else:
                return jsonify({'error': 'Fayl topilmadi'}), 404
        
        except Exception as e:
            logger.error(f"Delete file error: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @ocr_bp.route('/export-results/<format>', methods=['POST'])
    @login_required
    def export_results(format):
        """
        Ajratilgan ma'lumotlarni Excel yoki JSON qilib export qilish
        
        Args:
            format: 'excel' yoki 'json'
        
        Request body:
        {
            "results": [...]  # OCR natijalar
        }
        """
        try:
            if format not in ['excel', 'json']:
                return jsonify({'error': 'Noto\'g\'ri format'}), 400
            
            data = request.get_json()
            results = data.get('results', [])
            
            if format == 'json':
                # JSON export
                output_file = f"exports/ocr_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                os.makedirs('exports', exist_ok=True)
                
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(results, f, ensure_ascii=False, indent=2)
                
                return send_file(output_file, as_attachment=True, download_name=f"ocr_results.json")
            
            elif format == 'excel':
                # Excel export
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "OCR Results"
                
                # Headers
                headers = ['Invoice Number', 'Date', 'Customer', 'Total Amount', 'Items Count', 'Confidence']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                # Data
                for row_idx, result in enumerate(results, 2):
                    if result.get('success'):
                        data = result.get('data', {})
                        ws.cell(row=row_idx, column=1).value = data.get('invoice_number', 'N/A')
                        ws.cell(row=row_idx, column=2).value = data.get('invoice_date', 'N/A')
                        ws.cell(row=row_idx, column=3).value = data.get('customer_name', 'N/A')
                        ws.cell(row=row_idx, column=4).value = data.get('total_amount', 0)
                        ws.cell(row=row_idx, column=5).value = len(data.get('items', []))
                        ws.cell(row=row_idx, column=6).value = f"{data.get('confidence', 0) * 100:.1f}%"
                
                # Column widths
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                
                output_file = f"exports/ocr_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                os.makedirs('exports', exist_ok=True)
                wb.save(output_file)
                
                return send_file(output_file, as_attachment=True, download_name="ocr_results.xlsx")
        
        except Exception as e:
            logger.error(f"Export error: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    app.register_blueprint(ocr_bp)
    return ocr_bp
