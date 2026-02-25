"""
Report & Export Routes - Excel, PDF, Charts
"""
from flask import Blueprint, request, jsonify, send_file
from app.extensions import db
from app.models import Report, SalesOrder, Product, CashRegister, Expense
from app.main import login_required
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import csv

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/api/reports/sales-excel', methods=['GET'])
@login_required
def export_sales_excel():
    """Sotuvlar hisobotini Excel qilib export qilish"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Ma'lumotlarni filtrlash
    query = SalesOrder.query
    if start_date:
        query = query.filter(SalesOrder.order_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(SalesOrder.order_date <= datetime.fromisoformat(end_date))
    
    orders = query.all()
    
    # Excel workbook yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Sotuvlar"
    
    # Headers
    headers = ['ID', 'Raqam', 'Mijoz', 'Sana', 'Summa', 'Chegirma', 'Vergi', 'Holat']
    ws.append(headers)
    
    # Header formatini qo'lsh
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Ma'lumotlarni qo'shish
    for order in orders:
        ws.append([
            order.id,
            order.number,
            order.customer.name,
            order.order_date.strftime('%Y-%m-%d'),
            order.total_amount,
            order.discount,
            order.tax_amount,
            order.status
        ])
    
    # Ustunlar kengligini sozlash
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    # File ga saqlash
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'sotuvlar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@reports_bp.route('/api/reports/inventory-excel', methods=['GET'])
@login_required
def export_inventory_excel():
    """Inventar hisobotini Excel qilib export qilish"""
    products = Product.query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventar"
    
    headers = ['ID', 'Kodi', 'Nomi', 'Kategoriya', 'Birligi', 'Sotib olish narxi', 
               'Sotish narxi', 'Qoldiq', 'Minimal qoldiq']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for product in products:
        ws.append([
            product.id,
            product.code,
            product.name,
            product.category,
            product.unit,
            product.purchase_price,
            product.sale_price,
            product.quantity,
            product.minimum_quantity
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@reports_bp.route('/api/reports/cash-flow-chart', methods=['GET'])
@login_required
def cash_flow_chart():
    """Kassa oqimini diagramma shaklida"""
    days = request.args.get('days', 30, type=int)
    start_date = datetime.utcnow() - timedelta(days=days)
    
    # Ro'yxat uchun ma'lumotlarni to'plash
    date_range = {}
    for i in range(days):
        date = (start_date + timedelta(days=i)).date()
        date_range[str(date)] = {'in': 0, 'out': 0}
    
    # Haqiqiy ma'lumotlar kiritiladi
    cash_registers = CashRegister.query.all()
    
    data = {
        'labels': list(date_range.keys()),
        'datasets': [
            {
                'label': 'Kirim',
                'data': [date_range[d]['in'] for d in date_range],
                'borderColor': '#4472C4'
            },
            {
                'label': 'Chiqim',
                'data': [date_range[d]['out'] for d in date_range],
                'borderColor': '#ED7D31'
            }
        ]
    }
    
    return jsonify(data), 200


@reports_bp.route('/api/reports/sales-chart', methods=['GET'])
@login_required
def sales_chart():
    """Sotuvlar diagrammasi"""
    days = request.args.get('days', 30, type=int)
    start_date = datetime.utcnow() - timedelta(days=days)
    
    orders = SalesOrder.query.filter(SalesOrder.order_date >= start_date).all()
    
    # Har kuni sotuvlarni to'plash
    daily_sales = {}
    for order in orders:
        date_str = order.order_date.strftime('%Y-%m-%d')
        if date_str not in daily_sales:
            daily_sales[date_str] = 0
        daily_sales[date_str] += order.total_amount
    
    data = {
        'labels': list(daily_sales.keys()),
        'data': list(daily_sales.values()),
        'borderColor': '#70AD47',
        'backgroundColor': 'rgba(112, 173, 71, 0.1)'
    }
    
    return jsonify(data), 200


@reports_bp.route('/api/reports/summary', methods=['GET'])
@login_required
def summary_report():
    """Umumiy hisobot"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = SalesOrder.query
    if start_date:
        query = query.filter(SalesOrder.order_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(SalesOrder.order_date <= datetime.fromisoformat(end_date))
    
    orders = query.all()
    
    total_sales = sum([o.total_amount for o in orders])
    total_discount = sum([o.discount for o in orders])
    total_tax = sum([o.tax_amount for o in orders])
    
    expenses = Expense.query.all()
    total_expenses = sum([e.amount for e in expenses])
    
    profit = total_sales - total_expenses
    
    return jsonify({
        'period': f"{start_date} to {end_date}",
        'total_sales': total_sales,
        'total_discount': total_discount,
        'total_tax': total_tax,
        'total_expenses': total_expenses,
        'net_profit': profit,
        'orders_count': len(orders),
        'average_order': total_sales / len(orders) if orders else 0
    }), 200


@reports_bp.route('/api/reports/csv-export', methods=['GET'])
@login_required
def csv_export():
    """CSV formatda export"""
    report_type = request.args.get('type', 'sales')
    
    output = StringIO()
    writer = csv.writer(output)
    
    if report_type == 'sales':
        orders = SalesOrder.query.all()
        writer.writerow(['ID', 'Raqam', 'Mijoz', 'Sana', 'Summa', 'Holat'])
        for order in orders:
            writer.writerow([
                order.id,
                order.number,
                order.customer.name,
                order.order_date.strftime('%Y-%m-%d'),
                order.total_amount,
                order.status
            ])
    elif report_type == 'products':
        products = Product.query.all()
        writer.writerow(['ID', 'Kodi', 'Nomi', 'Narxi', 'Qoldiq'])
        for product in products:
            writer.writerow([
                product.id,
                product.code,
                product.name,
                product.sale_price,
                product.quantity
            ])
    
    output.seek(0)
    return send_file(
        BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'{report_type}_{datetime.now().strftime("%Y%m%d")}.csv'
    )
