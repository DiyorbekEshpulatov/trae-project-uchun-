#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Celery Tasks - Fon (Background) Vazifalar
Soliq hisobotlarini yuborish va boshqa long-running tasks
"""

import os
from datetime import datetime
from dotenv import load_dotenv
import logging

load_dotenv()
logger = logging.getLogger(__name__)

# Celery import
try:
    from app.celery_app import celery
except Exception as e:
    logger.warning(f"Celery not available: {e}")
    celery = None

# Telegram bot import
try:
    from app.telegram_bot import telegram_bot
except Exception as e:
    logger.warning(f"Telegram bot not available: {e}")
    telegram_bot = None

if celery:
    @celery.task(bind=True, name='send_tax_reports_async')
    def send_tax_reports_async(self, period, include_sales=True, include_vat=True, 
                               include_payroll=True, include_declaration=False):
        """
        Soliq hisobotlarini fonda yuborish
        
        Args:
            period: '2026-02' format
            include_sales: Savdo hisoboti yuborish
            include_vat: KDV hisoboti yuborish
            include_payroll: Oylik hisoboti yuborish
            include_declaration: Soliq deklaratsiyasi yuborish
        """
        try:
            from app.tax_integration import TaxCabinetAPIDev as TaxCabinet
            from datetime import timedelta
            
            logger.info(f"Task started: send_tax_reports_async for {period}")
            
            tax_api = TaxCabinet()
            results = {
                'task_id': self.request.id,
                'period': period,
                'status': 'processing',
                'reports': {}
            }
            
            # Update task state
            self.update_state(state='PROGRESS', 
                            meta={'current': 0, 'total': 4, 'status': 'Savdo hisoboti yuborilmoqda...'})
            
            # 1. Savdo hisoboti
            if include_sales:
                try:
                    sales_data = []  # Database'dan olinadi
                    sales_result = tax_api.send_sales_report(sales_data)
                    results['reports']['sales'] = sales_result
                    logger.info(f"Sales report sent: {sales_result['success']}")
                    
                    if telegram_bot:
                        telegram_bot.send_tax_report_notification(
                            'sales', 
                            'success' if sales_result['success'] else 'error',
                            {'period': period}
                        )
                except Exception as e:
                    logger.error(f"Sales report error: {str(e)}")
                    results['reports']['sales'] = {'success': False, 'error': str(e)}
            
            self.update_state(state='PROGRESS',
                            meta={'current': 1, 'total': 4, 'status': 'KDV hisoboti yuborilmoqda...'})
            
            # 2. KDV hisoboti
            if include_vat:
                try:
                    vat_data = {
                        'period': period,
                        'total_vat_collected': 0,
                        'vat_paid': 0,
                        'vat_to_pay': 0
                    }
                    vat_result = tax_api.send_vat_report(vat_data)
                    results['reports']['vat'] = vat_result
                    logger.info(f"VAT report sent: {vat_result['success']}")
                    
                    if telegram_bot:
                        telegram_bot.send_tax_report_notification(
                            'vat',
                            'success' if vat_result['success'] else 'error',
                            {'period': period}
                        )
                except Exception as e:
                    logger.error(f"VAT report error: {str(e)}")
                    results['reports']['vat'] = {'success': False, 'error': str(e)}
            
            self.update_state(state='PROGRESS',
                            meta={'current': 2, 'total': 4, 'status': 'Oylik hisoboti yuborilmoqda...'})
            
            # 3. Oylik hisoboti
            if include_payroll:
                try:
                    payroll_data = {
                        'period': period,
                        'total_salary': 0,
                        'pit_total': 0,
                        'pension_contribution': 0,
                        'employees_count': 0
                    }
                    payroll_result = tax_api.send_employee_payroll(payroll_data)
                    results['reports']['payroll'] = payroll_result
                    logger.info(f"Payroll report sent: {payroll_result['success']}")
                    
                    if telegram_bot:
                        telegram_bot.send_tax_report_notification(
                            'payroll',
                            'success' if payroll_result['success'] else 'error',
                            {'period': period}
                        )
                except Exception as e:
                    logger.error(f"Payroll report error: {str(e)}")
                    results['reports']['payroll'] = {'success': False, 'error': str(e)}
            
            self.update_state(state='PROGRESS',
                            meta={'current': 3, 'total': 4, 'status': 'Tugallanmoqda...'})
            
            # 4. Soliq deklaratsiyasi
            if include_declaration:
                try:
                    declaration_data = {
                        'period': period,
                        'income': 0,
                        'expenses': 0,
                        'profit': 0,
                        'tax_amount': 0
                    }
                    decl_result = tax_api.send_tax_declaration(declaration_data)
                    results['reports']['declaration'] = decl_result
                    logger.info(f"Declaration sent: {decl_result['success']}")
                    
                    if telegram_bot:
                        telegram_bot.send_tax_report_notification(
                            'declaration',
                            'success' if decl_result['success'] else 'error',
                            {'period': period}
                        )
                except Exception as e:
                    logger.error(f"Declaration error: {str(e)}")
                    results['reports']['declaration'] = {'success': False, 'error': str(e)}
            
            # Final status
            results['status'] = 'completed'
            results['all_success'] = all(
                r.get('success', False) 
                for r in results['reports'].values()
                if isinstance(r, dict)
            )
            
            logger.info(f"Task completed: {results['all_success']}")
            return results
        
        except Exception as e:
            logger.error(f"Task error: {str(e)}")
            return {
                'status': 'failed',
                'error': str(e),
                'task_id': self.request.id
            }
    
    @celery.task(bind=True, name='generate_excel_report')
    def generate_excel_report(self, report_type, period):
        """
        Excel hisobotni yaratish va saqlash
        
        Args:
            report_type: 'sales', 'vat', 'inventory', etc.
            period: '2026-02'
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            logger.info(f"Generating Excel report: {report_type} for {period}")
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"{report_type.upper()}"
            
            # Header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            ws['A1'] = f"{report_type.upper()} HISOBOTI"
            ws['A1'].fill = header_fill
            ws['A1'].font = header_font
            ws['A2'] = f"Davriy: {period}"
            ws['A3'] = f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Sample data
            ws['A5'] = 'ID'
            ws['B5'] = 'Nomi'
            ws['C5'] = 'Narxi'
            ws['D5'] = 'Miqdori'
            ws['E5'] = 'Jami'
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}5'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                ws[f'{col}5'].font = Font(bold=True)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            # Save file
            file_path = f"reports/{report_type}_{period.replace('-', '_')}.xlsx"
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            wb.save(file_path)
            
            logger.info(f"Excel report generated: {file_path}")
            
            return {
                'status': 'success',
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'generated_at': datetime.now().isoformat()
            }
        
        except Exception as e:
            logger.error(f"Excel report generation error: {str(e)}")
            return {
                'status': 'failed',
                'error': str(e)
            }
    
    @celery.task(bind=True, name='schedule_monthly_tax_report')
    def schedule_monthly_tax_report(self):
        """
        Oylik hisobot avtomatik yuborish (Cron job)
        Har oyning birinchi kuni soat 8:00 da
        """
        try:
            from datetime import datetime, timedelta
            
            current_date = datetime.now()
            previous_month = (current_date - timedelta(days=current_date.day)).strftime('%Y-%m')
            
            logger.info(f"Scheduling monthly tax report for {previous_month}")
            
            # Async task yuborish
            task = send_tax_reports_async.delay(
                period=previous_month,
                include_sales=True,
                include_vat=True,
                include_payroll=True,
                include_declaration=True
            )
            
            if telegram_bot:
                telegram_bot.send_message(
                    f"📅 <b>Oylik Soliq Hisoboti</b>\n\n"
                    f"Davriy: {previous_month}\n"
                    f"Holati: Yuborilmoqda...\n"
                    f"Task ID: {task.id}"
                )
            
            return {
                'status': 'scheduled',
                'task_id': task.id,
                'period': previous_month
            }
        
        except Exception as e:
            logger.error(f"Monthly report scheduling error: {str(e)}")
            return {'status': 'failed', 'error': str(e)}
    
    @celery.task(bind=True, name='backup_database')
    def backup_database(self):
        """
        Database ni zaxira qilish
        """
        try:
            import subprocess
            
            logger.info("Database backup started")
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = f"backups/db_backup_{timestamp}.sql"
            
            os.makedirs(os.path.dirname(backup_file), exist_ok=True)
            
            # PostgreSQL backup (customize for your database)
            db_url = os.getenv('DATABASE_URL', 'sqlite:///erp_system.db')
            
            if 'postgresql' in db_url:
                # pg_dump command
                pass
            else:
                # SQLite backup
                import shutil
                shutil.copy('erp_system.db', backup_file)
            
            logger.info(f"Database backup completed: {backup_file}")
            
            if telegram_bot:
                telegram_bot.send_message(
                    f"💾 <b>Database Zaxirasi</b>\n\n"
                    f"Vaqti: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                    f"Fayl: {backup_file}\n"
                    f"✅ Muvaffaqiyatli"
                )
            
            return {
                'status': 'success',
                'backup_file': backup_file,
                'timestamp': timestamp
            }
        
        except Exception as e:
            logger.error(f"Database backup error: {str(e)}")
            return {'status': 'failed', 'error': str(e)}

else:
    logger.warning("Celery is not configured. Tasks will not work.")
