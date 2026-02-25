#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Soliq integratsiyasi uchun routes moduli
Flask blueprint sifatida
"""

from flask import Blueprint, request, jsonify, session, render_template
from datetime import datetime, timedelta
from functools import wraps
import os
from dotenv import load_dotenv
import logging

load_dotenv()
logger = logging.getLogger(__name__)

# Tax API - environment-ga qarab real yoki dev version
USE_DEV_TAX_API = os.getenv('USE_DEV_TAX_API', 'True').lower() == 'true'

if USE_DEV_TAX_API:
    from app.tax_integration import TaxCabinetAPIDev as TaxCabinet
else:
    from app.tax_integration import TaxCabinetAPI as TaxCabinet

# Telegram Bot import
try:
    from app.telegram_bot import telegram_bot
except Exception as e:
    logger.warning(f"Telegram bot import failed: {e}")
    telegram_bot = None

tax_api = TaxCabinet()

def init_tax_routes(app, db, User, SalesOrder, Customer, Invoice, login_required, Report):
    """Tax routes-ni qayd qilish"""
    tax_bp = Blueprint('tax', __name__, url_prefix='/api/tax')
    
    @tax_bp.route('/send-all-reports', methods=['POST'])
    @login_required
    def send_all_reports():
        """
        Bir tugma orqali BARCHA hisobotlarni soliq kabenitiga yuborish
        
        Request body:
        {
            "period": "2026-02",
            "include_sales": true,
            "include_vat": true,
            "include_payroll": true
        }
        """
        try:
            user = db.session.query(User).filter_by(id=session['user_id']).first()
            if not user or user.role.name != 'admin':
                return jsonify({'error': 'Admin huquqi kerak'}), 403
            
            data = request.get_json()
            period = data.get('period', datetime.now().strftime('%Y-%m'))
            
            # Celery task orqali asinxron yuborish
            try:
                from app.celery_tasks import send_tax_reports_async
                
                task = send_tax_reports_async.delay(
                    period=period,
                    include_sales=data.get('include_sales', True),
                    include_vat=data.get('include_vat', True),
                    include_payroll=data.get('include_payroll', True),
                    include_declaration=data.get('include_declaration', False)
                )
                
                return jsonify({
                    'success': True,
                    'message': 'Hisobotlarni yuborish fonda boshlandi',
                    'task_id': task.id,
                    'period': period
                }), 202
        
        except Exception as e:
            logger.error(f"Barcha hisobotlarni yuborishda general xato: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @tax_bp.route('/tax-status', methods=['GET'])
    @login_required
    def get_tax_status():
        """Soliq kabenitidagi holati tekshirish"""
        try:
            status = tax_api.get_tax_status()
            return jsonify(status), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @tax_bp.route('/generate-reports/<report_type>', methods=['POST'])
    @login_required
    def generate_report(report_type):
        """
        Hisobot yaratish (Excel/PDF)
        report_type: sales, vat, payroll, declaration
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            
            user = db.session.query(User).filter_by(id=session['user_id']).first()
            if not user or user.role.name != 'admin':
                return jsonify({'error': 'Admin huquqi kerak'}), 403
            
            data = request.get_json()
            period = data.get('period', datetime.now().strftime('%Y-%m'))
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"{report_type.upper()} Report"
            
            # Header
            ws['A1'] = f"{report_type.upper()} HISOBOTI"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Davriy: {period}"
            ws['A3'] = f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Content - report_type ga qarab
            if report_type == 'sales':
                sales_orders = db.session.query(SalesOrder)\
                    .filter(SalesOrder.created_at >= f'{period}-01')\
                    .all()
                
                ws['A5'] = 'Sanasi'
                ws['B5'] = 'Mijoz'
                ws['C5'] = 'Mahsulot'
                ws['D5'] = 'Miqdori'
                ws['E5'] = 'Narxi'
                ws['F5'] = 'Jami'
                
                row = 6
                for order in sales_orders:
                    ws[f'A{row}'] = order.created_at.strftime('%Y-%m-%d')
                    ws[f'B{row}'] = order.customer.name if order.customer else 'N/A'
                    ws[f'C{row}'] = ', '.join([item.get('product_name', '') for item in (order.items or [])])
                    ws[f'D{row}'] = sum([item.get('quantity', 0) for item in (order.items or [])])
                    ws[f'E{row}'] = order.total_amount or 0
                    row += 1
            
            # File sifatida yuborish
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return {
                'success': True,
                'file_type': 'xlsx',
                'period': period,
                'generated_at': datetime.now().isoformat()
            }, 200
        
        except Exception as e:
            logger.error(f"Hisobot yaratishda xato: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @tax_bp.route('/send-telegram-notification', methods=['POST'])
    @login_required
    def send_telegram_notification():
        """Telegram orqali soliq hisoboti haqida xabar yuborish"""
        try:
            user = db.session.query(User).filter_by(id=session['user_id']).first()
            if not user or user.role.name not in ['admin', 'manager']:
                return jsonify({'error': 'Huquq kerak'}), 403
            
            data = request.get_json()
            message = data.get('message', 'Soliq hisobot yuborildi')
            
            # Telegram Bot API
            telegram_token = os.getenv('TELEGRAM_BOT_TOKEN', '')
            telegram_chat_id = os.getenv('TELEGRAM_CHAT_ID', '')
            
            if telegram_token and telegram_chat_id:
                telegram_url = f'https://api.telegram.org/bot{telegram_token}/sendMessage'
                telegram_data = {
                    'chat_id': telegram_chat_id,
                    'text': f"📊 {message}\n\nVaqti: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                }
                
                import requests
                response = requests.post(telegram_url, json=telegram_data)
                return jsonify({
                    'success': response.status_code == 200,
                    'message': 'Telegram xabari yuborildi'
                }), 200
            else:
                return jsonify({'error': 'Telegram bot konfiguratsiyasi etishmaydi'}), 400
        
        except Exception as e:
            logger.error(f"Telegram xabari yuborishda xato: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @tax_bp.route('/history', methods=['GET'])
    @login_required
    def get_tax_history():
        """Yuborilgan soliq hisobotlari tarixi"""
        try:
            page = request.args.get('page', 1, type=int)
            per_page = request.args.get('per_page', 20, type=int)
            
            # Filter for tax reports
            query = Report.query.filter(Report.report_type.like('tax_%')).order_by(Report.created_at.desc())
            
            pagination = query.paginate(page=page, per_page=per_page, error_out=False)
            reports = pagination.items
            
            return jsonify({
                'success': True,
                'reports': [{
                    'id': r.id,
                    'title': r.title,
                    'type': r.report_type,
                    'created_at': r.created_at.isoformat(),
                    'created_by': r.created_by_id,
                    'data': r.data
                } for r in reports],
                'total': pagination.total,
                'pages': pagination.pages,
                'current_page': page
            }), 200
        except Exception as e:
            logger.error(f"History error: {str(e)}")
            return jsonify({'error': str(e)}), 500

    @app.route('/tax-history')
    @login_required
    def tax_history_page():
        return render_template('tax_history.html')

    app.register_blueprint(tax_bp)
    return tax_bp
