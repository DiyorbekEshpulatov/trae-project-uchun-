#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Soliq integratsiyasi uchun routes moduli
Flask blueprint sifatida
"""

from flask import Blueprint, request, jsonify, session
from datetime import datetime, timedelta
from functools import wraps
import os
from dotenv import load_dotenv
import logging

load_dotenv()
logger = logging.getLogger(__name__)

# Tax API - environment-ga qarab real yoki dev version
USE_DEV_TAX_API = os.getenv('USE_DEV_TAX_API', 'True').lower() == 'true'

if USE_DEV_TAX_API:
    from app.tax_integration import TaxCabinetAPIDev as TaxCabinet
else:
    from app.tax_integration import TaxCabinetAPI as TaxCabinet

# Telegram Bot import
try:
    from app.telegram_bot import telegram_bot
except Exception as e:
    logger.warning(f"Telegram bot import failed: {e}")
    telegram_bot = None

tax_api = TaxCabinet()

def init_tax_routes(app, db, User, SalesOrder, Customer, Invoice, login_required):
    """Tax routes-ni qayd qilish"""
    tax_bp = Blueprint('tax', __name__, url_prefix='/api/tax')
    
    @tax_bp.route('/send-all-reports', methods=['POST'])
    @login_required
    def send_all_reports():
        """
        Bir tugma orqali BARCHA hisobotlarni soliq kabenitiga yuborish
        
        Request body:
        {
            "period": "2026-02",
            "include_sales": true,
            "include_vat": true,
            "include_payroll": true
        }
        """
        try:
            user = db.session.query(User).filter_by(id=session['user_id']).first()
            if not user or user.role.name != 'admin':
                return jsonify({'error': 'Admin huquqi kerak'}), 403
            
            data = request.get_json()
            period = data.get('period', datetime.now().strftime('%Y-%m'))
            
            results = {
                'timestamp': datetime.now().isoformat(),
                'period': period,
                'reports': {}
            }
            
            # 1. SAVDO HISOBOTI
            if data.get('include_sales', True):
                try:
                    sales_orders = db.session.query(SalesOrder)\
                        .filter(SalesOrder.created_at >= f'{period}-01')\
                        .all()
                    
                    sales_data = []
                    for order in sales_orders:
                        customer = order.customer
                        items = order.items or []
                        
                        sales_data.append({
                            'date': order.created_at.strftime('%Y-%m-%d'),
                            'customer_name': customer.name if customer else 'Unknown',
                            'customer_tin': customer.inn if customer else '',
                            'products': [
                                {
                                    'code': item.get('product_code', ''),
                                    'quantity': item.get('quantity', 0),
                                    'price': item.get('unit_price', 0),
                                    'amount': item.get('amount', 0)
                                }
                                for item in items
                            ],
                            'total_amount': order.total_amount or 0,
                            'vat_amount': (order.total_amount or 0) * 0.10  # 10% VAT
                        })
                    
                    sales_result = tax_api.send_sales_report(sales_data)
                    results['reports']['sales'] = sales_result
                    logger.info(f"Savdo hisoboti yuborildi: {sales_result['success']}")
                except Exception as e:
                    logger.error(f"Savdo hisoboti yuborishda xato: {str(e)}")
                    results['reports']['sales'] = {'success': False, 'error': str(e)}
            
            # 2. KDV (VAT) HISOBOTI
            if data.get('include_vat', True):
                try:
                    start_date = datetime.strptime(f'{period}-01', '%Y-%m-%d')
                    if period.endswith('-12'):
                        end_date = datetime(int(period[:4]) + 1, 1, 1) - timedelta(days=1)
                    else:
                        next_month = int(period.split('-')[1]) + 1
                        end_date = datetime(int(period[:4]), next_month, 1) - timedelta(days=1)
                    
                    invoices = db.session.query(Invoice)\
                        .filter(Invoice.created_at.between(start_date, end_date))\
                        .all()
                    
                    total_vat_collected = sum((inv.total_amount or 0) * 0.10 for inv in invoices)
                    
                    vat_data = {
                        'period': period,
                        'total_vat_collected': total_vat_collected,
                        'vat_paid': 0,
                        'vat_to_pay': total_vat_collected,
                        'invoice_count': len(invoices)
                    }
                    
                    vat_result = tax_api.send_vat_report(vat_data)
                    results['reports']['vat'] = vat_result
                    logger.info(f"KDV hisoboti yuborildi: {vat_result['success']}")
                except Exception as e:
                    logger.error(f"KDV hisoboti yuborishda xato: {str(e)}")
                    results['reports']['vat'] = {'success': False, 'error': str(e)}
            
            # 3. OYLIK HISOBOTI (Payroll)
            if data.get('include_payroll', True):
                try:
                    # Employees jadvali bo'lsa, oylik hisoboti yuborish
                    # Bu yerda sample data bilan
                    payroll_data = {
                        'period': period,
                        'total_salary': 0,
                        'pit_total': 0,
                        'pension_contribution': 0,
                        'employees_count': 0
                    }
                    
                    payroll_result = tax_api.send_employee_payroll(payroll_data)
                    results['reports']['payroll'] = payroll_result
                    logger.info(f"Oylik hisoboti yuborildi: {payroll_result['success']}")
                except Exception as e:
                    logger.error(f"Oylik hisoboti yuborishda xato: {str(e)}")
                    results['reports']['payroll'] = {'success': False, 'error': str(e)}
            
            # 4. SOLIQ DEKLARATSIYASI
            if data.get('include_declaration', False):
                try:
                    sales_orders = db.session.query(SalesOrder)\
                        .filter(SalesOrder.created_at >= f'{period}-01')\
                        .all()
                    
                    total_income = sum(order.total_amount or 0 for order in sales_orders)
                    estimated_expenses = total_income * 0.4
                    profit = total_income - estimated_expenses
                    tax_amount = profit * 0.12  # 12% corporate tax
                    
                    declaration_data = {
                        'period': period,
                        'income': total_income,
                        'expenses': estimated_expenses,
                        'profit': profit,
                        'tax_amount': tax_amount
                    }
                    
                    decl_result = tax_api.send_tax_declaration(declaration_data)
                    results['reports']['declaration'] = decl_result
                    logger.info(f"Deklaratsiya yuborildi: {decl_result['success']}")
                except Exception as e:
                    logger.error(f"Deklaratsiya yuborishda xato: {str(e)}")
                    results['reports']['declaration'] = {'success': False, 'error': str(e)}
            
            # Umumiy natija
            results['all_success'] = all(
                r.get('success', False) 
                for r in results['reports'].values() 
                if isinstance(r, dict)
            )
            
            # TELEGRAM NOTIFIKATSIYASI
            if telegram_bot and results['all_success']:
                telegram_bot.send_tax_summary({
                    'period': period,
                    'total_sales': sum(r.get('total_amount', 0) for r in (results.get('sales_data', []))),
                    'vat_amount': results.get('reports', {}).get('vat', {}).get('data', {}).get('total_vat_collected', 0),
                    'profit': 0,
                    'tax_amount': 0
                })
            
            return jsonify(results), 200
        
        except Exception as e:
            logger.error(f"Barcha hisobotlarni yuborishda general xato: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @tax_bp.route('/tax-status', methods=['GET'])
    @login_required
    def get_tax_status():
        """Soliq kabenitidagi holati tekshirish"""
        try:
            status = tax_api.get_tax_status()
            return jsonify(status), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @tax_bp.route('/generate-reports/<report_type>', methods=['POST'])
    @login_required
    def generate_report(report_type):
        """
        Hisobot yaratish (Excel/PDF)
        report_type: sales, vat, payroll, declaration
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            
            user = db.session.query(User).filter_by(id=session['user_id']).first()
            if not user or user.role.name != 'admin':
                return jsonify({'error': 'Admin huquqi kerak'}), 403
            
            data = request.get_json()
            period = data.get('period', datetime.now().strftime('%Y-%m'))
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"{report_type.upper()} Report"
            
            # Header
            ws['A1'] = f"{report_type.upper()} HISOBOTI"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Davriy: {period}"
            ws['A3'] = f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Content - report_type ga qarab
            if report_type == 'sales':
                sales_orders = db.session.query(SalesOrder)\
                    .filter(SalesOrder.created_at >= f'{period}-01')\
                    .all()
                
                ws['A5'] = 'Sanasi'
                ws['B5'] = 'Mijoz'
                ws['C5'] = 'Mahsulot'
                ws['D5'] = 'Miqdori'
                ws['E5'] = 'Narxi'
                ws['F5'] = 'Jami'
                
                row = 6
                for order in sales_orders:
                    ws[f'A{row}'] = order.created_at.strftime('%Y-%m-%d')
                    ws[f'B{row}'] = order.customer.name if order.customer else 'N/A'
                    ws[f'C{row}'] = ', '.join([item.get('product_name', '') for item in (order.items or [])])
                    ws[f'D{row}'] = sum([item.get('quantity', 0) for item in (order.items or [])])
                    ws[f'E{row}'] = order.total_amount or 0
                    row += 1
            
            # File sifatida yuborish
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return {
                'success': True,
                'file_type': 'xlsx',
                'period': period,
                'generated_at': datetime.now().isoformat()
            }, 200
        
        except Exception as e:
            logger.error(f"Hisobot yaratishda xato: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    @tax_bp.route('/send-telegram-notification', methods=['POST'])
    @login_required
    def send_telegram_notification():
        """Telegram orqali soliq hisoboti haqida xabar yuborish"""
        try:
            user = db.session.query(User).filter_by(id=session['user_id']).first()
            if not user or user.role.name not in ['admin', 'manager']:
                return jsonify({'error': 'Huquq kerak'}), 403
            
            data = request.get_json()
            message = data.get('message', 'Soliq hisobot yuborildi')
            
            # Telegram Bot API
            telegram_token = os.getenv('TELEGRAM_BOT_TOKEN', '')
            telegram_chat_id = os.getenv('TELEGRAM_CHAT_ID', '')
            
            if telegram_token and telegram_chat_id:
                telegram_url = f'https://api.telegram.org/bot{telegram_token}/sendMessage'
                telegram_data = {
                    'chat_id': telegram_chat_id,
                    'text': f"📊 {message}\n\nVaqti: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                }
                
                import requests
                response = requests.post(telegram_url, json=telegram_data)
                return jsonify({
                    'success': response.status_code == 200,
                    'message': 'Telegram xabari yuborildi'
                }), 200
            else:
                return jsonify({'error': 'Telegram bot konfiguratsiyasi etishmaydi'}), 400
        
        except Exception as e:
            logger.error(f"Telegram xabari yuborishda xato: {str(e)}")
            return jsonify({'error': str(e)}), 500
    
    app.register_blueprint(tax_bp)
    return tax_bp
