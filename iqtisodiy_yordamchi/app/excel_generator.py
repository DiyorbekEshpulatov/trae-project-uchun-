#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Jadvallar va Formulalar Yaratish Moduli
Savdo, Sotib olish, Inventar, Moliyaviy hisobotlari
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from typing import Dict, List, Optional

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelTableGenerator:
    """Excel jadvallarni yaratish va formulalar qo'shish"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        self.total_font = Font(bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_sales_table(self, sales_data: List[Dict], output_file: str = 'sales_report.xlsx'):
        """
        Savdo jadvalini yaratish
        
        sales_data: [{
            'date': '2026-02-04',
            'order_id': 'ORD-001',
            'customer': 'ABC Corp',
            'product': 'Product 1',
            'quantity': 10,
            'unit_price': 100000,
            'discount': 5,  # %
            'total': 950000
        }]
        """
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Sales"
            
            # Headers
            headers = ['Sanasi', 'Buyurtma ID', 'Mijoz', 'Mahsulot', 'Miqdori', 'Narxi', 'Chegirma %', 'Jami']
            for col, header in enumerate(headers, 1):
                cell = self.ws.cell(row=1, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
            
            # Data rows
            for idx, row_data in enumerate(sales_data, 2):
                self.ws.cell(row=idx, column=1).value = row_data.get('date')
                self.ws.cell(row=idx, column=2).value = row_data.get('order_id')
                self.ws.cell(row=idx, column=3).value = row_data.get('customer')
                self.ws.cell(row=idx, column=4).value = row_data.get('product')
                self.ws.cell(row=idx, column=5).value = row_data.get('quantity', 0)
                self.ws.cell(row=idx, column=6).value = row_data.get('unit_price', 0)
                self.ws.cell(row=idx, column=7).value = row_data.get('discount', 0)
                
                # Formula: Jami = Miqdori * Narxi * (1 - Chegirma/100)
                self.ws.cell(row=idx, column=8).value = f"=E{idx}*F{idx}*(1-G{idx}/100)"
                
                # Format numbers
                for col in [5, 6, 7, 8]:
                    cell = self.ws.cell(row=idx, column=col)
                    if col in [6, 8]:
                        cell.number_format = '#,##0'
                    elif col == 7:
                        cell.number_format = '0.0'
                    cell.border = self.border
            
            # Summary row
            total_row = len(sales_data) + 3
            self.ws.cell(row=total_row, column=3).value = "JAMI:"
            self.ws.cell(row=total_row, column=3).font = self.total_font
            self.ws.cell(row=total_row, column=3).fill = self.total_fill
            
            # Total formula
            self.ws.cell(row=total_row, column=8).value = f"=SUM(H2:H{len(sales_data)+1})"
            self.ws.cell(row=total_row, column=8).font = self.total_font
            self.ws.cell(row=total_row, column=8).fill = self.total_fill
            self.ws.cell(row=total_row, column=8).number_format = '#,##0'
            
            # Column widths
            widths = [15, 15, 20, 20, 10, 15, 12, 15]
            for idx, width in enumerate(widths, 1):
                self.ws.column_dimensions[get_column_letter(idx)].width = width
            
            self.wb.save(output_file)
            logger.info(f"Savdo jadvali yaratildi: {output_file}")
            
            return {
                'success': True,
                'file': output_file,
                'rows': len(sales_data),
                'message': 'Savdo jadvali muvaffaqiyatli yaratildi'
            }
        except Exception as e:
            logger.error(f"Savdo jadvali yaratishda xato: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def create_purchase_table(self, purchase_data: List[Dict], output_file: str = 'purchase_report.xlsx'):
        """Sotib olish jadvalini yaratish"""
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Purchases"
            
            # Headers
            headers = ['Sanasi', 'PO ID', 'Etkazuvchi', 'Mahsulot', 'Miqdori', 'Birlik Narxi', 'Jami']
            for col, header in enumerate(headers, 1):
                cell = self.ws.cell(row=1, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            # Data rows
            for idx, row_data in enumerate(purchase_data, 2):
                self.ws.cell(row=idx, column=1).value = row_data.get('date')
                self.ws.cell(row=idx, column=2).value = row_data.get('po_id')
                self.ws.cell(row=idx, column=3).value = row_data.get('supplier')
                self.ws.cell(row=idx, column=4).value = row_data.get('product')
                self.ws.cell(row=idx, column=5).value = row_data.get('quantity', 0)
                self.ws.cell(row=idx, column=6).value = row_data.get('unit_price', 0)
                
                # Formula: Jami = Miqdori * Narxi
                self.ws.cell(row=idx, column=7).value = f"=E{idx}*F{idx}"
                
                for col in [5, 6, 7]:
                    cell = self.ws.cell(row=idx, column=col)
                    cell.number_format = '#,##0'
                    cell.border = self.border
            
            # Summary row
            total_row = len(purchase_data) + 3
            self.ws.cell(row=total_row, column=3).value = "JAMI:"
            self.ws.cell(row=total_row, column=3).font = self.total_font
            self.ws.cell(row=total_row, column=3).fill = self.total_fill
            
            self.ws.cell(row=total_row, column=7).value = f"=SUM(G2:G{len(purchase_data)+1})"
            self.ws.cell(row=total_row, column=7).font = self.total_font
            self.ws.cell(row=total_row, column=7).fill = self.total_fill
            self.ws.cell(row=total_row, column=7).number_format = '#,##0'
            
            # Column widths
            for idx, width in enumerate([15, 15, 20, 20, 10, 15, 15], 1):
                self.ws.column_dimensions[get_column_letter(idx)].width = width
            
            self.wb.save(output_file)
            logger.info(f"Sotib olish jadvali yaratildi: {output_file}")
            
            return {
                'success': True,
                'file': output_file,
                'rows': len(purchase_data),
                'message': 'Sotib olish jadvali muvaffaqiyatli yaratildi'
            }
        except Exception as e:
            logger.error(f"Sotib olish jadvali yaratishda xato: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def create_inventory_table(self, inventory_data: List[Dict], output_file: str = 'inventory_report.xlsx'):
        """
        Inventar jadvalini yaratish
        
        inventory_data: [{
            'code': 'P001',
            'name': 'Mahsulot 1',
            'category': 'Kategoriya',
            'opening_stock': 100,
            'purchase': 50,
            'sales': 30,
            'closing_stock': 120,
            'unit_cost': 100000
        }]
        """
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Inventory"
            
            # Headers
            headers = ['Kod', 'Nomi', 'Kategoriya', 'Kirim', 'Sotib Olish', 'Savdo', 'Qolgan', 'Birlik Narxi', 'Jami Narx']
            for col, header in enumerate(headers, 1):
                cell = self.ws.cell(row=1, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            # Data rows
            for idx, row_data in enumerate(inventory_data, 2):
                self.ws.cell(row=idx, column=1).value = row_data.get('code')
                self.ws.cell(row=idx, column=2).value = row_data.get('name')
                self.ws.cell(row=idx, column=3).value = row_data.get('category')
                self.ws.cell(row=idx, column=4).value = row_data.get('opening_stock', 0)
                self.ws.cell(row=idx, column=5).value = row_data.get('purchase', 0)
                self.ws.cell(row=idx, column=6).value = row_data.get('sales', 0)
                
                # Formula: Qolgan = Kirim + Sotib Olish - Savdo
                self.ws.cell(row=idx, column=7).value = f"=D{idx}+E{idx}-F{idx}"
                self.ws.cell(row=idx, column=8).value = row_data.get('unit_cost', 0)
                
                # Formula: Jami Narx = Qolgan * Birlik Narxi
                self.ws.cell(row=idx, column=9).value = f"=G{idx}*H{idx}"
                
                for col in [4, 5, 6, 7, 8, 9]:
                    cell = self.ws.cell(row=idx, column=col)
                    cell.number_format = '#,##0'
                    cell.border = self.border
            
            # Summary rows
            total_row = len(inventory_data) + 3
            self.ws.cell(row=total_row, column=2).value = "JAMI:"
            self.ws.cell(row=total_row, column=2).font = self.total_font
            self.ws.cell(row=total_row, column=2).fill = self.total_fill
            
            for col in [7, 9]:
                cell = self.ws.cell(row=total_row, column=col)
                cell.value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{len(inventory_data)+1})"
                cell.font = self.total_font
                cell.fill = self.total_fill
                cell.number_format = '#,##0'
            
            # Column widths
            widths = [10, 20, 15, 10, 15, 10, 10, 15, 15]
            for idx, width in enumerate(widths, 1):
                self.ws.column_dimensions[get_column_letter(idx)].width = width
            
            self.wb.save(output_file)
            logger.info(f"Inventar jadvali yaratildi: {output_file}")
            
            return {
                'success': True,
                'file': output_file,
                'rows': len(inventory_data),
                'message': 'Inventar jadvali muvaffaqiyatli yaratildi'
            }
        except Exception as e:
            logger.error(f"Inventar jadvali yaratishda xato: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def create_financial_report(self, financial_data: Dict, output_file: str = 'financial_report.xlsx'):
        """
        Moliyaviy hisobot yaratish
        
        financial_data: {
            'income': 5000000,
            'expenses': 2000000,
            'purchases': 1500000,
            'salary': 500000,
            'utilities': 0
        }
        """
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Financial Report"
            
            # Title
            self.ws.merge_cells('A1:D1')
            title = self.ws['A1']
            title.value = "MOLIYAVIY HISOBOT"
            title.font = Font(bold=True, size=14, color="FFFFFF")
            title.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
            title.alignment = Alignment(horizontal='center', vertical='center')
            
            # Period
            self.ws['A2'] = f"Davriy: {datetime.now().strftime('%Y-%m-%d')}"
            
            # Income section
            self.ws['A4'] = "DAROMAD"
            self.ws['A4'].font = Font(bold=True, size=12)
            self.ws['A4'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            self.ws['A5'] = "Savdo Daromadi"
            self.ws['B5'] = financial_data.get('income', 0)
            self.ws['B5'].number_format = '#,##0'
            
            # Expenses section
            self.ws['A7'] = "XARAJATLAR"
            self.ws['A7'].font = Font(bold=True, size=12)
            self.ws['A7'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
            self.ws['A8'] = "Sotib Olish"
            self.ws['B8'] = financial_data.get('purchases', 0)
            self.ws['B8'].number_format = '#,##0'
            
            self.ws['A9'] = "Oylik Xarajatlari"
            self.ws['B9'] = financial_data.get('salary', 0)
            self.ws['B9'].number_format = '#,##0'
            
            self.ws['A10'] = "Boshqa Xarajatlar"
            self.ws['B10'] = financial_data.get('expenses', 0) - financial_data.get('purchases', 0) - financial_data.get('salary', 0)
            self.ws['B10'].number_format = '#,##0'
            
            # Total Expenses
            self.ws['A11'] = "JAMI XARAJATLAR"
            self.ws['A11'].font = Font(bold=True)
            self.ws['A11'].fill = self.total_fill
            self.ws['B11'] = f"=B8+B9+B10"
            self.ws['B11'].font = Font(bold=True)
            self.ws['B11'].fill = self.total_fill
            self.ws['B11'].number_format = '#,##0'
            
            # Profit/Loss
            self.ws['A13'] = "FOYDA/ZARAR"
            self.ws['A13'].font = Font(bold=True, size=12)
            self.ws['A13'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            self.ws['B13'] = f"=B5-B11"
            self.ws['B13'].font = Font(bold=True, size=12)
            self.ws['B13'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            self.ws['B13'].number_format = '#,##0'
            
            # Percentages
            self.ws['A15'] = "FOIZLAR"
            self.ws['A15'].font = Font(bold=True, size=12)
            
            self.ws['A16'] = "Foyda Foizi"
            self.ws['B16'] = f"=IF(B5=0,0,(B13/B5)*100)"
            self.ws['B16'].number_format = '0.00"%"'
            
            self.ws['A17'] = "Xarajat Foizi"
            self.ws['B17'] = f"=IF(B5=0,0,(B11/B5)*100)"
            self.ws['B17'].number_format = '0.00"%"'
            
            # Column widths
            self.ws.column_dimensions['A'].width = 25
            self.ws.column_dimensions['B'].width = 20
            
            self.wb.save(output_file)
            logger.info(f"Moliyaviy hisobot yaratildi: {output_file}")
            
            return {
                'success': True,
                'file': output_file,
                'message': 'Moliyaviy hisobot muvaffaqiyatli yaratildi'
            }
        except Exception as e:
            logger.error(f"Moliyaviy hisobot yaratishda xato: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def create_account_ledger(self, ledger_data: List[Dict], account_name: str, output_file: str = 'ledger.xlsx'):
        """
        Hisobi ledger yaratish
        
        ledger_data: [{
            'date': '2026-02-04',
            'description': 'Savdo',
            'debit': 500000,
            'credit': 0,
            'balance': 500000
        }]
        """
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = account_name
            
            # Header
            self.ws.merge_cells('A1:E1')
            title = self.ws['A1']
            title.value = f"HISOBI LEDGER - {account_name}"
            title.font = Font(bold=True, size=12, color="FFFFFF")
            title.fill = self.header_fill
            title.alignment = Alignment(horizontal='center')
            
            # Column headers
            headers = ['Sanasi', 'Tafsifoti', 'Debet', 'Kredit', 'Balansi']
            for col, header in enumerate(headers, 1):
                cell = self.ws.cell(row=3, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            
            # Data rows
            balance = 0
            for idx, row_data in enumerate(ledger_data, 4):
                self.ws.cell(row=idx, column=1).value = row_data.get('date')
                self.ws.cell(row=idx, column=2).value = row_data.get('description')
                self.ws.cell(row=idx, column=3).value = row_data.get('debit', 0)
                self.ws.cell(row=idx, column=4).value = row_data.get('credit', 0)
                
                # Formula: Balance = Previous Balance + Debit - Credit
                self.ws.cell(row=idx, column=5).value = f"=E{idx-1}+C{idx}-D{idx}"
                
                for col in [3, 4, 5]:
                    cell = self.ws.cell(row=idx, column=col)
                    cell.number_format = '#,##0'
                    cell.border = self.border
            
            # Summary
            total_row = len(ledger_data) + 5
            self.ws.cell(row=total_row, column=2).value = "JAMI"
            self.ws.cell(row=total_row, column=2).font = self.total_font
            self.ws.cell(row=total_row, column=2).fill = self.total_fill
            
            for col in [3, 4]:
                cell = self.ws.cell(row=total_row, column=col)
                cell.value = f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{len(ledger_data)+3})"
                cell.font = self.total_font
                cell.fill = self.total_fill
                cell.number_format = '#,##0'
            
            # Column widths
            widths = [15, 25, 15, 15, 15]
            for idx, width in enumerate(widths, 1):
                self.ws.column_dimensions[get_column_letter(idx)].width = width
            
            self.wb.save(output_file)
            logger.info(f"Ledger yaratildi: {output_file}")
            
            return {
                'success': True,
                'file': output_file,
                'rows': len(ledger_data),
                'message': 'Ledger muvaffaqiyatli yaratildi'
            }
        except Exception as e:
            logger.error(f"Ledger yaratishda xato: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def create_multi_sheet_workbook(self, data_dict: Dict, output_file: str = 'complete_report.xlsx'):
        """
        Ko'p sheet'li workbook yaratish (Savdo, Sotib olish, Inventar, Moliyaviy)
        """
        try:
            self.wb = Workbook()
            self.wb.remove(self.wb.active)  # Default sheet'ni o'chirish
            
            # Sales sheet
            if 'sales' in data_dict:
                self.ws = self.wb.create_sheet("Sales")
                self._create_sales_sheet(data_dict['sales'])
            
            # Purchase sheet
            if 'purchases' in data_dict:
                self.ws = self.wb.create_sheet("Purchases")
                self._create_purchase_sheet(data_dict['purchases'])
            
            # Inventory sheet
            if 'inventory' in data_dict:
                self.ws = self.wb.create_sheet("Inventory")
                self._create_inventory_sheet(data_dict['inventory'])
            
            # Financial sheet
            if 'financial' in data_dict:
                self.ws = self.wb.create_sheet("Financial")
                self._create_financial_sheet(data_dict['financial'])
            
            # Summary sheet
            self.ws = self.wb.create_sheet("Summary", 0)
            self._create_summary_sheet(data_dict)
            
            self.wb.save(output_file)
            logger.info(f"Multi-sheet workbook yaratildi: {output_file}")
            
            return {
                'success': True,
                'file': output_file,
                'sheets': list(self.wb.sheetnames),
                'message': 'To\'liq hisobot muvaffaqiyatli yaratildi'
            }
        except Exception as e:
            logger.error(f"Multi-sheet workbook yaratishda xato: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def _create_summary_sheet(self, data_dict: Dict):
        """Xulosa sheet'i yaratish"""
        self.ws['A1'] = "UMUMIY XULOSA"
        self.ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        self.ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        self.ws.merge_cells('A1:B1')
        
        row = 3
        
        if 'sales' in data_dict:
            total_sales = sum(item.get('total', 0) for item in data_dict['sales'])
            self.ws[f'A{row}'] = "Jami Savdo"
            self.ws[f'B{row}'] = total_sales
            self.ws[f'B{row}'].number_format = '#,##0'
            row += 1
        
        if 'purchases' in data_dict:
            total_purchases = sum(item.get('total', 0) for item in data_dict['purchases'])
            self.ws[f'A{row}'] = "Jami Sotib Olish"
            self.ws[f'B{row}'] = total_purchases
            self.ws[f'B{row}'].number_format = '#,##0'
            row += 1
        
        if 'financial' in data_dict:
            self.ws[f'A{row}'] = "Foyda"
            self.ws[f'B{row}'] = data_dict['financial'].get('profit', 0)
            self.ws[f'B{row}'].number_format = '#,##0'
    
    def _create_sales_sheet(self, sales_data: List[Dict]):
        """Sales sheet'i yaratish"""
        headers = ['Sanasi', 'Buyurtma', 'Mijoz', 'Mahsulot', 'Miqdori', 'Narxi', 'Jami']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        for idx, row_data in enumerate(sales_data, 2):
            self.ws.cell(row=idx, column=1).value = row_data.get('date')
            self.ws.cell(row=idx, column=2).value = row_data.get('order_id')
            self.ws.cell(row=idx, column=3).value = row_data.get('customer')
            self.ws.cell(row=idx, column=4).value = row_data.get('product')
            self.ws.cell(row=idx, column=5).value = row_data.get('quantity')
            self.ws.cell(row=idx, column=6).value = row_data.get('unit_price')
            self.ws.cell(row=idx, column=7).value = row_data.get('total')
            self.ws.cell(row=idx, column=7).number_format = '#,##0'
    
    def _create_purchase_sheet(self, purchase_data: List[Dict]):
        """Purchase sheet'i yaratish"""
        headers = ['Sanasi', 'PO ID', 'Etkazuvchi', 'Mahsulot', 'Miqdori', 'Narxi', 'Jami']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        for idx, row_data in enumerate(purchase_data, 2):
            self.ws.cell(row=idx, column=1).value = row_data.get('date')
            self.ws.cell(row=idx, column=2).value = row_data.get('po_id')
            self.ws.cell(row=idx, column=3).value = row_data.get('supplier')
            self.ws.cell(row=idx, column=4).value = row_data.get('product')
            self.ws.cell(row=idx, column=5).value = row_data.get('quantity')
            self.ws.cell(row=idx, column=6).value = row_data.get('unit_price')
            self.ws.cell(row=idx, column=7).value = row_data.get('total')
            self.ws.cell(row=idx, column=7).number_format = '#,##0'
    
    def _create_inventory_sheet(self, inventory_data: List[Dict]):
        """Inventory sheet'i yaratish"""
        headers = ['Kod', 'Nomi', 'Kategoriya', 'Kirim', 'Sotib Olish', 'Savdo', 'Qolgan', 'Narxi', 'Jami']
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        for idx, row_data in enumerate(inventory_data, 2):
            self.ws.cell(row=idx, column=1).value = row_data.get('code')
            self.ws.cell(row=idx, column=2).value = row_data.get('name')
            self.ws.cell(row=idx, column=3).value = row_data.get('category')
            self.ws.cell(row=idx, column=4).value = row_data.get('opening_stock')
            self.ws.cell(row=idx, column=5).value = row_data.get('purchase')
            self.ws.cell(row=idx, column=6).value = row_data.get('sales')
            self.ws.cell(row=idx, column=7).value = f"=D{idx}+E{idx}-F{idx}"
            self.ws.cell(row=idx, column=8).value = row_data.get('unit_cost')
            self.ws.cell(row=idx, column=9).value = f"=G{idx}*H{idx}"
            self.ws.cell(row=idx, column=7).number_format = '#,##0'
            self.ws.cell(row=idx, column=8).number_format = '#,##0'
            self.ws.cell(row=idx, column=9).number_format = '#,##0'
    
    def _create_financial_sheet(self, financial_data: Dict):
        """Financial sheet'i yaratish"""
        self.ws['A1'] = "MOLIYAVIY HISOBOT"
        self.ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        self.ws['A1'].fill = self.header_fill
        
        self.ws['A3'] = "Daromad"
        self.ws['B3'] = financial_data.get('income', 0)
        self.ws['B3'].number_format = '#,##0'
        
        self.ws['A4'] = "Sotib Olish"
        self.ws['B4'] = financial_data.get('purchases', 0)
        self.ws['B4'].number_format = '#,##0'
        
        self.ws['A5'] = "Boshqa Xarajatlar"
        self.ws['B5'] = financial_data.get('expenses', 0)
        self.ws['B5'].number_format = '#,##0'
        
        self.ws['A7'] = "Jami Xarajatlar"
        self.ws['A7'].font = Font(bold=True)
        self.ws['B7'] = "=B4+B5"
        self.ws['B7'].font = Font(bold=True)
        self.ws['B7'].number_format = '#,##0'
        
        self.ws['A9'] = "FOYDA"
        self.ws['A9'].font = Font(bold=True, size=12)
        self.ws['A9'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.ws['B9'] = "=B3-B7"
        self.ws['B9'].font = Font(bold=True, size=12)
        self.ws['B9'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.ws['B9'].number_format = '#,##0'
